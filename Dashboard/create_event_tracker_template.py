"""
Create Excel template for event tracker format
"""
import pandas as pd
from pathlib import Path

def create_event_tracker_template(output_path: str = "../templates/Event_Tracker_Template.xlsx"):
    """Create Excel template for event-by-event tracking"""
    
    # Create directory if it doesn't exist
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Individual Events Sheet - with example data structure
    individual_events_columns = [
        'Set',           # 1, 2, 3, etc.
        'Point',         # 1, 2, 3, etc. (point number within set)
        'Rotation',      # 1-6 (rotation number)
        'Player',        # Player name
        'Position',      # OH1, OH2, MB1, MB2, OPP, S, L
        'Action',        # serve, receive, set, attack, block, dig
        'Outcome',       # Varies by action (see config.py for valid outcomes)
        'Attack_Type',   # normal, tip, after_block (required only for attacks)
        'Notes'          # Optional notes
    ]
    
    # Create example rows for Individual Events
    example_individual_events = [
        {
            'Set': 1,
            'Point': 1,
            'Rotation': 1,
            'Player': 'Alex',
            'Position': 'OH1',
            'Action': 'serve',
            'Outcome': 'good',
            'Attack_Type': '',
            'Notes': 'Example serve'
        },
        {
            'Set': 1,
            'Point': 1,
            'Rotation': 1,
            'Player': 'John',
            'Position': 'L',
            'Action': 'receive',
            'Outcome': 'perfect_0',
            'Attack_Type': '',
            'Notes': 'Perfect reception within 1m'
        },
        {
            'Set': 1,
            'Point': 1,
            'Rotation': 1,
            'Player': 'Mike',
            'Position': 'S',
            'Action': 'set',
            'Outcome': 'exceptional',
            'Attack_Type': '',
            'Notes': 'Exceptional set'
        },
        {
            'Set': 1,
            'Point': 1,
            'Rotation': 1,
            'Player': 'Alex',
            'Position': 'OH1',
            'Action': 'attack',
            'Outcome': 'kill',
            'Attack_Type': 'normal',
            'Notes': 'Attack kill - normal attack'
        },
        {
            'Set': 1,
            'Point': 2,
            'Rotation': 1,
            'Player': 'Sarah',
            'Position': 'MB1',
            'Action': 'attack',
            'Outcome': 'blocked',
            'Attack_Type': 'tip',
            'Notes': 'Tip attack blocked'
        },
        {
            'Set': 1,
            'Point': 3,
            'Rotation': 2,
            'Player': 'David',
            'Position': 'MB2',
            'Action': 'block',
            'Outcome': 'touch',
            'Attack_Type': '',
            'Notes': 'Block touch'
        },
        {
            'Set': 1,
            'Point': 4,
            'Rotation': 2,
            'Player': 'Emma',
            'Position': 'L',
            'Action': 'dig',
            'Outcome': 'good_1',
            'Attack_Type': '',
            'Notes': 'Good dig within 3m'
        }
    ]
    
    df_individual = pd.DataFrame(example_individual_events)
    df_individual = df_individual.reindex(columns=individual_events_columns)
    
    # Team Events Sheet - with example data structure
    team_events_columns = [
        'Set',           # 1, 2, 3, etc.
        'Point',         # 1, 2, 3, etc.
        'Rotation',      # 1-6
        'Point_Type',    # serving or receiving
        'Point Won',     # yes or no
        'Our_Score',     # Optional: auto-calculated
        'Opponent_Score', # Optional: auto-calculated
        'Rally_Length'   # Optional: number of actions in rally
    ]
    
    # Create example rows for Team Events
    example_team_events = [
        {
            'Set': 1,
            'Point': 1,
            'Rotation': 1,
            'Point_Type': 'serving',
            'Point Won': 'yes',
            'Our_Score': 1,
            'Opponent_Score': 0,
            'Rally_Length': 4
        },
        {
            'Set': 1,
            'Point': 2,
            'Rotation': 1,
            'Point_Type': 'serving',
            'Point Won': 'no',
            'Our_Score': 1,
            'Opponent_Score': 1,
            'Rally_Length': 5
        },
        {
            'Set': 1,
            'Point': 3,
            'Rotation': 2,
            'Point_Type': 'receiving',
            'Point Won': 'yes',
            'Our_Score': 2,
            'Opponent_Score': 1,
            'Rally_Length': 6
        },
        {
            'Set': 1,
            'Point': 4,
            'Rotation': 2,
            'Point_Type': 'receiving',
            'Point Won': 'no',
            'Our_Score': 2,
            'Opponent_Score': 2,
            'Rally_Length': 3
        }
    ]
    
    df_team = pd.DataFrame(example_team_events)
    df_team = df_team.reindex(columns=team_events_columns)
    
    # Create Excel file with both sheets
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_individual.to_excel(writer, sheet_name='Individual Events', index=False)
        df_team.to_excel(writer, sheet_name='Team Events', index=False)
        
        # Get workbook and worksheet objects for formatting
        workbook = writer.book
        ws_individual = writer.sheets['Individual Events']
        ws_team = writer.sheets['Team Events']
        
        # Format headers (bold)
        from openpyxl.styles import Font, PatternFill
        
        header_fill = PatternFill(start_color="040C7B", end_color="040C7B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for ws in [ws_individual, ws_team]:
            for cell in ws[1]:  # Header row
                cell.font = header_font
                cell.fill = header_fill
        
        # Auto-adjust column widths
        for ws in [ws_individual, ws_team]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Event Tracker Template created at: {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    output = sys.argv[1] if len(sys.argv) > 1 else "../templates/Event_Tracker_Template.xlsx"
    create_event_tracker_template(output)

