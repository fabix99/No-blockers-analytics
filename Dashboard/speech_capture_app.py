"""
Speech-to-Data Capture System for Volleyball Match Analysis
Voice-controlled action capture with UI configuration for set, rotation, serving status
"""
import streamlit as st
try:
    import speech_recognition as sr
    SPEECH_AVAILABLE = True
except ImportError:
    SPEECH_AVAILABLE = False
    st.error("⚠️ SpeechRecognition not installed. Run: pip install SpeechRecognition")

try:
    import pyaudio
    AUDIO_AVAILABLE = True
except ImportError:
    AUDIO_AVAILABLE = False
    st.warning("⚠️ PyAudio not installed. Microphone access will not work. To fix: Install portaudio first (brew install portaudio), then pip install pyaudio")

import threading
import time
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
from utils import sanitize_template_path

# Initialize session state
if 'current_set' not in st.session_state:
    st.session_state.current_set = 1
if 'player_positions' not in st.session_state:
    st.session_state.player_positions = {}  # {player_name: position}
if 'current_rotation' not in st.session_state:
    st.session_state.current_rotation = 1  # Setter position (1-6)
if 'is_serving' not in st.session_state:
    st.session_state.is_serving = True
if 'current_server' not in st.session_state:
    st.session_state.current_server = None  # Selected server player name
if 'is_recording' not in st.session_state:
    st.session_state.is_recording = False
if 'recognition_text' not in st.session_state:
    st.session_state.recognition_text = ""
if 'transcript_history' not in st.session_state:
    st.session_state.transcript_history = []  # List of all recognized text
if 'mic_status' not in st.session_state:
    st.session_state.mic_status = "Not started"
if 'mic_error' not in st.session_state:
    st.session_state.mic_error = None
if 'captured_actions' not in st.session_state:
    st.session_state.captured_actions = []  # Actions from current recording
if 'confirmed_points' not in st.session_state:
    st.session_state.confirmed_points = []  # All confirmed point data
if 'show_confirmation' not in st.session_state:
    st.session_state.show_confirmation = False

# Volleyball actions and outcomes
ACTIONS = {
    'attack': ['attack', 'hit', 'spike', 'swing'],
    'serve': ['serve', 'service', 'serving'],
    'block': ['block', 'blocking', 'blocked'],
    'receive': ['receive', 'reception', 'pass', 'dig'],
    'set': ['set', 'setting', 'setter']
}

OUTCOMES = {
    'kill': ['kill', 'point', 'scored', 'won'],
    'ace': ['ace', 'direct point'],
    'good': ['good', 'success', 'successful', 'positive'],
    'error': ['error', 'miss', 'missed', 'out', 'fault'],
    'touch': ['touch', 'touched', 'touch block']
}

# Volleyball position layout (counter-clockwise from serve position)
POSITION_LAYOUT = {
    1: "Right-Back (Server)",
    2: "Right-Front",
    3: "Middle-Front",
    4: "Left-Front",
    5: "Left-Back",
    6: "Middle-Back"
}

def parse_speech(text):
    """Parse speech text to extract only actions (action, player, outcome)"""
    text_lower = text.lower().strip()
    
    # Try to extract action
    action_found = None
    for action, keywords in ACTIONS.items():
        if any(kw in text_lower for kw in keywords):
            action_found = action
            break
    
    # Try to extract outcome
    outcome_found = None
    for outcome, keywords in OUTCOMES.items():
        if any(kw in text_lower for kw in keywords):
            outcome_found = outcome
            break
    
    # Extract player name (look for capitalized words, prefer known players)
    player_found = None
    words = text.split()
    
    # First try to match known players
    for word in words:
        if word in st.session_state.player_positions:
            player_found = word
            break
    
    # If not found, look for capitalized words
    if not player_found:
        for word in words:
            if word[0].isupper() and len(word) > 2:
                # Skip common volleyball terms
                if word.lower() not in ['attack', 'serve', 'block', 'set', 'receive', 'outside', 'middle', 'opposite', 'libero', 'rotation', 'point', 'won', 'lost']:
                    player_found = word
                    break
    
    if action_found and player_found:
        return {
            'action': action_found,
            'player': player_found,
            'outcome': outcome_found or 'good',
            'text': text
        }
    
    return None

def recognize_speech_continuous():
    """Continuous speech recognition in background thread"""
    if not SPEECH_AVAILABLE or not AUDIO_AVAILABLE:
        st.session_state.mic_error = "Speech recognition or audio not available. Please install dependencies."
        st.session_state.is_recording = False
        return
    
    recognizer = sr.Recognizer()
    recognizer.energy_threshold = 300  # Lower threshold to be more sensitive
    recognizer.dynamic_energy_threshold = True
    recognizer.pause_threshold = 0.8
    
    try:
        microphone = sr.Microphone()
        st.session_state.mic_error = None
        st.session_state.mic_status = "Microphone initialized"
    except Exception as e:
        st.session_state.mic_error = f"Microphone initialization error: {e}"
        st.session_state.mic_status = f"Error: {e}"
        st.session_state.is_recording = False
        return
    
    # Calibrate microphone
    try:
        with microphone as source:
            st.session_state.mic_status = "Calibrating microphone... (1 second)"
            recognizer.adjust_for_ambient_noise(source, duration=1)
            st.session_state.mic_status = "Microphone ready - listening..."
    except Exception as e:
        st.session_state.mic_error = f"Microphone calibration error: {e}"
        st.session_state.mic_status = f"Calibration failed: {e}"
        st.session_state.is_recording = False
        return
    
    st.session_state.mic_status = "🎤 Listening for speech..."
    listen_count = 0
    
    # Use microphone context manager properly - like the test button
    try:
        with microphone as source:
            while st.session_state.is_recording:
                try:
                    # Listen with longer timeout and phrase limit
                    listen_count += 1
                    st.session_state.mic_status = f"🎤 Listening... (attempt {listen_count})"
                    audio = recognizer.listen(source, timeout=2, phrase_time_limit=8)
                    
                    try:
                        # Use Google Speech Recognition (requires internet)
                        st.session_state.mic_status = "🔄 Processing speech..."
                        text = recognizer.recognize_google(audio)
                        
                        if text:
                            st.session_state.recognition_text = text
                            st.session_state.mic_status = f"✅ Heard: {text}"
                            
                            # Add to transcript history
                            st.session_state.transcript_history.append(text)
                            # Keep only last 20 transcript entries
                            if len(st.session_state.transcript_history) > 20:
                                st.session_state.transcript_history = st.session_state.transcript_history[-20:]
                            
                            # Parse and add action
                            parsed = parse_speech(text)
                            if parsed:
                                st.session_state.captured_actions.append(parsed)
                                st.session_state.mic_status = f"✅ Action captured: {parsed['action']} by {parsed['player']}"
                            else:
                                st.session_state.mic_status = f"⚠️ Heard but couldn't parse: {text}"
                            
                    except sr.UnknownValueError:
                        st.session_state.mic_status = "❓ Could not understand audio (try speaking more clearly)"
                        # Don't break, just continue listening
                        continue
                    except sr.RequestError as e:
                        st.session_state.mic_error = f"Speech recognition API error: {e}"
                        st.session_state.mic_status = f"❌ API Error: {e}"
                        # Break on API errors as they're persistent
                        break
                    
                except sr.WaitTimeoutError:
                    # This is normal - no speech detected, keep listening
                    continue
                except Exception as e:
                    st.session_state.mic_error = f"Unexpected error: {e}"
                    st.session_state.mic_status = f"❌ Error: {e}"
                    break
    except Exception as e:
        st.session_state.mic_error = f"Microphone context error: {e}"
        st.session_state.mic_status = f"❌ Context Error: {e}"
    
    st.session_state.mic_status = "⏹️ Recording stopped"

def write_to_excel(template_path, output_path, points_data):
    """Write captured data to Excel template format"""
    try:
        # Load template or create new workbook
        try:
            wb = load_workbook(template_path)
        except:
            # Create new workbook if template doesn't exist
            wb = load_workbook()
        
        # Group points by set and rotation
        by_set_rotation = {}
        team_points = {}
        
        for point in points_data:
            set_num = point['set']
            rot = point['rotation']
            serving = point['serving']
            
            if set_num not in by_set_rotation:
                by_set_rotation[set_num] = {}
            if rot not in by_set_rotation[set_num]:
                by_set_rotation[set_num][rot] = {
                    'players': {},
                    'serving': serving
                }
            
            # Aggregate actions by player
            for action_data in point['actions']:
                player = action_data['player']
                position = st.session_state.player_positions.get(player, '')
                
                if player not in by_set_rotation[set_num][rot]['players']:
                    by_set_rotation[set_num][rot]['players'][player] = {
                        'position': position,
                        'Attack_Kills': 0, 'Attack_Good': 0, 'Attack_Errors': 0,
                        'Service_Aces': 0, 'Service_Good': 0, 'Service_Errors': 0,
                        'Block_Kills': 0, 'Block_Touches': 0, 'Block_Errors': 0,
                        'Reception_Good': 0, 'Reception_Errors': 0,
                        'Sets_Exceptional': 0, 'Sets_Good': 0, 'Sets_Errors': 0
                    }
                
                # Update position if we have it
                if position:
                    by_set_rotation[set_num][rot]['players'][player]['position'] = position
                
                # Map action + outcome to stat column
                action = action_data['action']
                outcome = action_data['outcome']
                
                if action == 'attack':
                    if outcome == 'kill':
                        by_set_rotation[set_num][rot]['players'][player]['Attack_Kills'] += 1
                    elif outcome == 'good':
                        by_set_rotation[set_num][rot]['players'][player]['Attack_Good'] += 1
                    elif outcome == 'error':
                        by_set_rotation[set_num][rot]['players'][player]['Attack_Errors'] += 1
                
                elif action == 'serve':
                    if outcome == 'ace':
                        by_set_rotation[set_num][rot]['players'][player]['Service_Aces'] += 1
                    elif outcome == 'good':
                        by_set_rotation[set_num][rot]['players'][player]['Service_Good'] += 1
                    elif outcome == 'error':
                        by_set_rotation[set_num][rot]['players'][player]['Service_Errors'] += 1
                
                elif action == 'block':
                    if outcome == 'kill':
                        by_set_rotation[set_num][rot]['players'][player]['Block_Kills'] += 1
                    elif outcome == 'touch':
                        by_set_rotation[set_num][rot]['players'][player]['Block_Touches'] += 1
                    elif outcome == 'error':
                        by_set_rotation[set_num][rot]['players'][player]['Block_Errors'] += 1
                
                elif action == 'receive':
                    if outcome == 'good':
                        by_set_rotation[set_num][rot]['players'][player]['Reception_Good'] += 1
                    elif outcome == 'error':
                        by_set_rotation[set_num][rot]['players'][player]['Reception_Errors'] += 1
                
                elif action == 'set':
                    if outcome == 'good':
                        by_set_rotation[set_num][rot]['players'][player]['Sets_Good'] += 1
                    elif outcome == 'error':
                        by_set_rotation[set_num][rot]['players'][player]['Sets_Errors'] += 1
            
            # Team points
            if set_num not in team_points:
                team_points[set_num] = {
                    'Serving_Rallies': 0, 'Serving_Points_Won': 0, 'Serving_Points_Lost': 0,
                    'Receiving_Rallies': 0, 'Receiving_Points_Won': 0, 'Receiving_Points_Lost': 0
                }
            
            rallies = point.get('rallies', 1)
            result = point.get('result', 'won')
            
            if serving == 'serving':
                team_points[set_num]['Serving_Rallies'] += rallies
                if result == 'won':
                    team_points[set_num]['Serving_Points_Won'] += 1
                else:
                    team_points[set_num]['Serving_Points_Lost'] += 1
            else:
                team_points[set_num]['Receiving_Rallies'] += rallies
                if result == 'won':
                    team_points[set_num]['Receiving_Points_Won'] += 1
                else:
                    team_points[set_num]['Receiving_Points_Lost'] += 1
        
        # Write to sheets
        for set_num, rotations in by_set_rotation.items():
            for rot, data in rotations.items():
                sheet_name = f"Set{set_num}-Rot{rot}"
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                    # Add headers
                    stat_columns = ['Attack_Kills', 'Attack_Good', 'Attack_Errors',
                                   'Service_Aces', 'Service_Good', 'Service_Errors',
                                   'Block_Kills', 'Block_Touches', 'Block_Errors',
                                   'Reception_Good', 'Reception_Errors',
                                   'Sets_Exceptional', 'Sets_Good', 'Sets_Errors']
                    headers = ['Player', 'Position'] + stat_columns
                    ws.append(headers)
                else:
                    ws = wb[sheet_name]
                
                # Write player data
                for player, stats in data['players'].items():
                    position = stats.pop('position', '')
                    stat_values = [stats.get(col, 0) for col in stat_columns]
                    row = [player, position] + stat_values
                    ws.append(row)
            
            # Team points sheet
            team_sheet_name = f"Set{set_num}-Team Points"
            if team_sheet_name not in wb.sheetnames:
                ws_team = wb.create_sheet(team_sheet_name)
                ws_team.append(['Serving_Rallies', 'Serving_Points_Won', 'Serving_Points_Lost',
                               'Receiving_Rallies', 'Receiving_Points_Won', 'Receiving_Points_Lost'])
            else:
                ws_team = wb[team_sheet_name]
            
            ws_team.append(list(team_points[set_num].values()))
        
        # Save
        wb.save(output_path)
        return True
    
    except Exception as e:
        st.error(f"Error writing to Excel: {e}")
        import traceback
        st.code(traceback.format_exc())
        return False

def main():
    st.title("🎤 Speech-to-Data Match Capture")
    
    # ==================== TOP CONFIGURATION BAR (COMPACT) ====================
    # Check dependencies first
    if not SPEECH_AVAILABLE or not AUDIO_AVAILABLE:
        st.error("⚠️ **Speech recognition dependencies missing!**")
        st.markdown("""
        **To install:**
        1. Install Homebrew (if not installed): `/bin/bash -c "$(curl -fsSL https://raw.githubusercontent.com/Homebrew/install/HEAD/install.sh)"`
        2. Install portaudio: `brew install portaudio`
        3. Install Python packages: `pip install SpeechRecognition pyaudio`
        """)
        st.stop()
    
    # Compact top bar with all configuration
    config_row1_col1, config_row1_col2, config_row1_col3, config_row1_col4 = st.columns(4)
    with config_row1_col1:
        current_set = st.selectbox("**Set**", options=[1, 2, 3, 4, 5], 
                                   index=st.session_state.current_set - 1,
                                   key='set_selector')
        st.session_state.current_set = current_set
    with config_row1_col2:
        serving_status = st.radio("**Status**", ["Serving", "Receiving"], 
                                  index=0 if st.session_state.is_serving else 1,
                                  key='serving_radio', horizontal=True)
        st.session_state.is_serving = (serving_status == "Serving")
    with config_row1_col3:
        rotation_sel = st.selectbox("**Rotation** (Setter Pos)", 
                                    options=[1, 2, 3, 4, 5, 6],
                                    index=st.session_state.current_rotation - 1,
                                    key='rotation_selector',
                                    help=f"Setter position determines rotation. Current: {POSITION_LAYOUT[st.session_state.current_rotation]}")
        st.session_state.current_rotation = rotation_sel
    with config_row1_col4:
        st.markdown("<br>", unsafe_allow_html=True)  # Spacing
        if st.button("📖 Help", use_container_width=True):
            st.info("""
            **Quick Guide:**
            1. Fill player roster below
            2. Click Start Recording → Speak actions → Stop
            3. Confirm point with rallies & result
            4. Export when done
            """)
    
    # Second row: Server selector
    server_col1, server_col2, server_col3, server_col4 = st.columns(4)
    with server_col1:
        # Get list of players from roster
        available_players = list(st.session_state.player_positions.keys()) if st.session_state.player_positions else ["No players set"]
        server_options = ["None"] + available_players if available_players != ["No players set"] else ["None"]
        current_server_idx = 0
        if st.session_state.current_server in available_players:
            current_server_idx = available_players.index(st.session_state.current_server) + 1
        
        selected_server = st.selectbox("**Server**", 
                                       options=server_options,
                                       index=current_server_idx,
                                       key='server_selector',
                                       help="Select the player who is serving")
        st.session_state.current_server = selected_server if selected_server != "None" else None
        if st.session_state.current_server:
            st.caption(f"Current: {st.session_state.current_server}")
    with server_col2, server_col3, server_col4:
        pass  # Empty columns for spacing
    
    st.markdown("---")
    
    # ==================== PLAYER ROSTER (HORIZONTAL GRID) ====================
    st.markdown("#### 👥 Player Roster")
    
    # Find current players for each position
    position_players = {}
    for player, pos in st.session_state.player_positions.items():
        if pos not in position_players:
            position_players[pos] = []
        position_players[pos].append(player)
    
    # Horizontal grid for positions - 7 positions total
    pos_cols = st.columns(7)
    positions = ["OH1", "OH2", "MB1", "MB2", "S", "OPP", "L"]
    
    for idx, pos in enumerate(positions):
        with pos_cols[idx]:
            # Get the first player for this position (if multiple exist, just show first)
            current_players = position_players.get(pos, [])
            current_player = current_players[0] if current_players else ''
            
            player_name = st.text_input(f"**{pos}**", 
                                       value=current_player,
                                       key=f'pos_{pos}',
                                       placeholder="Player name",
                                       label_visibility="collapsed")
            st.caption(pos)
            
            if player_name:
                # Remove old player(s) from this position
                if pos in position_players:
                    for old_player in position_players[pos]:
                        if old_player in st.session_state.player_positions:
                            # Only remove if this position is still assigned to them
                            if st.session_state.player_positions.get(old_player) == pos:
                                del st.session_state.player_positions[old_player]
                # Add new mapping
                st.session_state.player_positions[player_name] = pos
            elif pos in position_players:
                # Clear position if input is empty
                for old_player in position_players[pos]:
                    if old_player in st.session_state.player_positions:
                        if st.session_state.player_positions.get(old_player) == pos:
                            del st.session_state.player_positions[old_player]
    
    st.markdown("---")
    
    # ==================== RECORDING INSTRUCTIONS ====================
    if st.session_state.is_recording:
        st.info("""
        🎤 **Speak in this format:** **[Action] [Player Name] [Outcome]**
        
        **Examples:**
        - "Attack Fabio Kill"
        - "Serve David Ace"
        - "Block Alex Good"
        - "Receive Maria Good"
        - "Set John Good"
        
        **Actions:** Attack, Serve, Block, Receive, Set  
        **Outcomes:** Kill, Ace, Good, Error, Touch
        """)
    else:
        st.caption("💡 **Tip:** When recording, say: **[Action] [Player] [Outcome]** (e.g., 'Attack Fabio Kill')")
    
    # ==================== RECORDING CONTROLS (COMPACT) ====================
    rec_col1, rec_col2, rec_col3, rec_col4 = st.columns([2, 2, 3, 3])
    
    with rec_col1:
        if st.button("▶️ Start Recording", disabled=st.session_state.is_recording, 
                    type="primary", use_container_width=True):
            st.session_state.is_recording = True
            st.session_state.captured_actions = []
            st.session_state.recognition_text = ""
            st.session_state.transcript_history = []  # Clear transcript on new recording
            st.session_state.show_confirmation = False
            thread = threading.Thread(target=recognize_speech_continuous, daemon=True)
            thread.start()
            st.rerun()
    
    with rec_col2:
        stop_clicked = st.button("⏹️ Stop Recording", disabled=not st.session_state.is_recording,
                    type="secondary", use_container_width=True)
        if stop_clicked:
            st.session_state.is_recording = False
            st.session_state.show_confirmation = True
            st.rerun()
    
    with rec_col3:
        if st.session_state.is_recording:
            status_text = st.session_state.get('mic_status', '🔴 RECORDING')
            st.markdown(f"🔴 **RECORDING**")
            if 'mic_status' in st.session_state and st.session_state.mic_status:
                st.caption(st.session_state.mic_status)
        else:
            if st.session_state.show_confirmation and st.session_state.captured_actions:
                st.markdown("✅ **Recording stopped** - Review below")
            else:
                st.markdown("⚪ Not Recording")
                if 'mic_error' in st.session_state and st.session_state.mic_error:
                    st.error(st.session_state.mic_error)
    
    with rec_col4:
        if st.session_state.recognition_text:
            st.caption(f"**Last heard:** {st.session_state.recognition_text[:50]}")
        if st.session_state.is_recording and st.session_state.captured_actions:
            st.caption(f"**Actions captured:** {len(st.session_state.captured_actions)}")
    
    # ==================== LIVE TRANSCRIPT ====================
    if st.session_state.is_recording or (st.session_state.transcript_history and st.session_state.show_confirmation):
        st.markdown("### 📝 Live Transcript")
        if st.session_state.transcript_history:
            # Show transcript in reverse order (most recent first)
            transcript_display = "\n".join([f"• {text}" for text in reversed(st.session_state.transcript_history[-10:])])
            st.text_area("Recognized Speech", value=transcript_display, height=150, 
                        disabled=True, key='transcript_display',
                        help="This shows what the speech recognition is hearing in real-time")
        else:
            status_msg = st.session_state.get('mic_status', 'Waiting for speech input...')
            st.info(f"🎤 {status_msg} Speak actions clearly: 'Attack Fabio Kill'")
        
        # Show microphone status
        if st.session_state.is_recording:
            if 'mic_error' in st.session_state and st.session_state.mic_error:
                st.error(f"❌ **Error:** {st.session_state.mic_error}")
            elif 'mic_status' in st.session_state:
                st.caption(f"Status: {st.session_state.mic_status}")
    
    # Test microphone button
    if not st.session_state.is_recording:
        test_col1, test_col2 = st.columns([3, 1])
        with test_col2:
            if st.button("🔍 Test Microphone", help="Test if microphone is accessible"):
                try:
                    import speech_recognition as sr
                    r = sr.Recognizer()
                    mic = sr.Microphone()
                    with mic as source:
                        st.info("🎤 Say something for 2 seconds...")
                        audio = r.listen(source, timeout=2, phrase_time_limit=2)
                    text = r.recognize_google(audio)
                    st.success(f"✅ Microphone works! Heard: '{text}'")
                except sr.UnknownValueError:
                    st.warning("⚠️ Microphone works but couldn't understand speech. Check volume/permissions.")
                except Exception as e:
                    st.error(f"❌ Microphone error: {e}")
    
    # Auto-refresh during recording to show live updates
    # Use a placeholder that updates without blocking
    if st.session_state.is_recording:
        # Check if we need to refresh (status changed)
        if 'last_mic_status' not in st.session_state:
            st.session_state.last_mic_status = ""
        
        # Only rerun if status actually changed to avoid unnecessary reruns
        current_status = st.session_state.get('mic_status', '')
        if current_status != st.session_state.last_mic_status:
            st.session_state.last_mic_status = current_status
        st.rerun()
    
    st.markdown("---")
    
    # ==================== CONFIRMATION SECTION (COMPACT) ====================
    if st.session_state.show_confirmation:
        st.markdown("---")
        if st.session_state.captured_actions:
            st.markdown("### ✅ Confirm Point Data")
            st.info(f"📊 **{len(st.session_state.captured_actions)} action(s) captured** - Review and confirm below")
            
            conf_col1, conf_col2 = st.columns([3, 1])
            
            with conf_col1:
                # Show captured actions in compact table
                actions_df = pd.DataFrame(st.session_state.captured_actions)
                st.dataframe(actions_df[['action', 'player', 'outcome']], use_container_width=True, hide_index=True)
                
                # Quick delete options
                if len(st.session_state.captured_actions) > 0:
                    st.caption("**Delete actions:**")
                    del_cols = st.columns(min(len(st.session_state.captured_actions), 6))
                    for i, action in enumerate(st.session_state.captured_actions[:6]):
                        with del_cols[i % len(del_cols)]:
                            if st.button("❌", key=f"del_{i}", help=f"Delete: {action['action']} {action['player']}"):
                                st.session_state.captured_actions.pop(i)
                                st.rerun()
            
            with conf_col2:
                st.markdown("**Point Details:**")
                num_rallies = st.number_input("**Rallies**", min_value=1, value=1, key='rallies_input')
                point_result = st.radio("**Result**", ["Won", "Lost"], key='result_radio')
                
                st.markdown("---")
                if st.button("✅ Confirm & Save Point", type="primary", use_container_width=True):
                    confirmed_point = {
                        'set': st.session_state.current_set,
                        'rotation': st.session_state.current_rotation,
                        'serving': 'serving' if st.session_state.is_serving else 'receiving',
                        'server': st.session_state.current_server,
                        'actions': st.session_state.captured_actions.copy(),
                        'rallies': num_rallies,
                        'result': point_result.lower()
                    }
                    st.session_state.confirmed_points.append(confirmed_point)
                    st.session_state.captured_actions = []
                    st.session_state.show_confirmation = False
                    st.session_state.recognition_text = ""
                    st.success(f"✅ Point #{len(st.session_state.confirmed_points)} saved!")
                    st.rerun()
        else:
            st.warning("⚠️ **No actions captured yet.** Click Start Recording and speak actions, then click Stop.")
            if st.button("🔄 Try Again", use_container_width=True):
                st.session_state.show_confirmation = False
                st.rerun()
        
        st.markdown("---")
    
    # ==================== CAPTURED POINTS & EXPORT (SIDE BY SIDE) ====================
    points_col, export_col = st.columns([2, 1])
    
    with points_col:
        st.markdown("### 📊 Captured Points")
        if st.session_state.confirmed_points:
            st.write(f"**{len(st.session_state.confirmed_points)} points captured**")
            summary_data = []
            for i, point in enumerate(st.session_state.confirmed_points, 1):
                summary_data.append({
                    '#': i,
                    'Set': point['set'],
                    'Rot': point['rotation'],
                    'Status': point['serving'][:4],
                    'Server': point.get('server', 'N/A') or 'N/A',
                    'Actions': len(point['actions']),
                    'Rallies': point['rallies'],
                    'Result': point['result']
                })
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
            
            # Compact delete option
            if st.checkbox("🗑️ Delete points", key='show_delete'):
                del_cols = st.columns(min(len(st.session_state.confirmed_points), 6))
                for i, point in enumerate(st.session_state.confirmed_points[:6]):  # Show first 6
                    with del_cols[i % len(del_cols)]:
                        if st.button(f"Del #{i+1}", key=f"del_point_{i}"):
                            st.session_state.confirmed_points.pop(i)
                            st.rerun()
        else:
            st.info("No points captured yet")
    
    with export_col:
        st.markdown("### 💾 Export")
        team_name = st.text_input("Team Name", value="", placeholder="No Blockers", key='team_name')
        game_date = st.date_input("Game Date", value=datetime.now().date(), key='game_date')
        template_path_input = st.text_input("Template", value="../templates/Match_Template.xlsx", key='template_path')
        
        # Validate and sanitize template path
        template_path = None
        if template_path_input:
            try:
                template_path = str(sanitize_template_path(template_path_input))
            except (ValueError, FileNotFoundError) as e:
                st.error(f"❌ Invalid template path: {e}")
                template_path = None
        
        if team_name:
            date_str = game_date.strftime('%Y%m%d')
            default_filename = f"{team_name}_{date_str}.xlsx"
        else:
            default_filename = f"match_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_filename = st.text_input("Filename", value=default_filename, key='output_filename')
        
        if st.button("📥 Export to Excel", type="primary", use_container_width=True):
            if not team_name:
                st.warning("⚠️ Enter team name")
            elif not template_path:
                st.warning("⚠️ Valid template path required")
            elif st.session_state.confirmed_points:
                if write_to_excel(template_path, output_filename, st.session_state.confirmed_points):
                    st.success(f"✅ Exported!")
                    with open(output_filename, 'rb') as f:
                        st.download_button("⬇️ Download", f.read(), output_filename, 
                                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.warning("No points to export")
    
    # Compact instructions at bottom
    with st.expander("📖 Quick Help"):
        help_col1, help_col2 = st.columns(2)
        with help_col1:
            st.markdown("""
            **Workflow:**
            1. Fill player roster
            2. Start Recording → Speak → Stop
            3. Confirm with rallies & result
            4. Export when done
            """)
        with help_col2:
            st.markdown("""
            **Voice Examples:**
            - "Attack Fabio Kill"
            - "Serve David Ace"
            - "Block Alex Good"
            
            **Actions:** Attack, Serve, Block, Receive, Set  
            **Outcomes:** Kill, Ace, Good, Error, Touch
            """)

if __name__ == "__main__":
    main()
