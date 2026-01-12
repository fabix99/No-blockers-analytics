"""
Create sample event tracker data for testing
"""
import pandas as pd
from pathlib import Path

def create_sample_event_data(output_path: str = "../data/examples/sample_event_tracker.xlsx"):
    """Create sample event tracker data for testing"""
    
    # Create directory if it doesn't exist
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Sample Individual Events
    individual_events = [
        # Set 1, Point 1 - Serving rally won
        {'Set': 1, 'Point': 1, 'Rotation': 1, 'Player': 'Alex', 'Position': 'OH1', 'Action': 'serve', 'Outcome': 'good', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 1, 'Rotation': 1, 'Player': 'John', 'Position': 'L', 'Action': 'receive', 'Outcome': 'perfect_0', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 1, 'Rotation': 1, 'Player': 'Mike', 'Position': 'S', 'Action': 'set', 'Outcome': 'exceptional', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 1, 'Rotation': 1, 'Player': 'Alex', 'Position': 'OH1', 'Action': 'attack', 'Outcome': 'kill', 'Attack_Type': 'normal', 'Notes': ''},
        
        # Set 1, Point 2 - Receiving rally lost
        {'Set': 1, 'Point': 2, 'Rotation': 1, 'Player': 'Sarah', 'Position': 'OH2', 'Action': 'receive', 'Outcome': 'poor_2', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 2, 'Rotation': 1, 'Player': 'Mike', 'Position': 'S', 'Action': 'set', 'Outcome': 'poor', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 2, 'Rotation': 1, 'Player': 'David', 'Position': 'MB1', 'Action': 'attack', 'Outcome': 'blocked', 'Attack_Type': 'normal', 'Notes': ''},
        
        # Set 1, Point 3 - Receiving rally won
        {'Set': 1, 'Point': 3, 'Rotation': 2, 'Player': 'John', 'Position': 'L', 'Action': 'receive', 'Outcome': 'good_1', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 3, 'Rotation': 2, 'Player': 'Mike', 'Position': 'S', 'Action': 'set', 'Outcome': 'good', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 3, 'Rotation': 2, 'Player': 'Sarah', 'Position': 'OH2', 'Action': 'attack', 'Outcome': 'kill', 'Attack_Type': 'tip', 'Notes': ''},
        
        # Set 1, Point 4 - Serving rally with block
        {'Set': 1, 'Point': 4, 'Rotation': 2, 'Player': 'David', 'Position': 'MB1', 'Action': 'serve', 'Outcome': 'good', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 4, 'Rotation': 2, 'Player': 'Emma', 'Position': 'MB2', 'Action': 'block', 'Outcome': 'touch', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 4, 'Rotation': 2, 'Player': 'Emma', 'Position': 'L', 'Action': 'dig', 'Outcome': 'perfect_0', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 4, 'Rotation': 2, 'Player': 'Mike', 'Position': 'S', 'Action': 'set', 'Outcome': 'good', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 4, 'Rotation': 2, 'Player': 'Alex', 'Position': 'OH1', 'Action': 'attack', 'Outcome': 'kill', 'Attack_Type': 'after_block', 'Notes': ''},
        
        # Set 1, Point 5 - Attack error
        {'Set': 1, 'Point': 5, 'Rotation': 3, 'Player': 'Sarah', 'Position': 'OH2', 'Action': 'serve', 'Outcome': 'ace', 'Attack_Type': '', 'Notes': ''},
        
        # Set 1, Point 6 - Attack defended
        {'Set': 1, 'Point': 6, 'Rotation': 3, 'Player': 'David', 'Position': 'MB1', 'Action': 'serve', 'Outcome': 'good', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 6, 'Rotation': 3, 'Player': 'John', 'Position': 'L', 'Action': 'receive', 'Outcome': 'good_1', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 6, 'Rotation': 3, 'Player': 'Mike', 'Position': 'S', 'Action': 'set', 'Outcome': 'exceptional', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 6, 'Rotation': 3, 'Player': 'Sarah', 'Position': 'OH2', 'Action': 'attack', 'Outcome': 'defended', 'Attack_Type': 'normal', 'Notes': ''},
        {'Set': 1, 'Point': 6, 'Rotation': 3, 'Player': 'Emma', 'Position': 'L', 'Action': 'dig', 'Outcome': 'good_1', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 6, 'Rotation': 3, 'Player': 'Mike', 'Position': 'S', 'Action': 'set', 'Outcome': 'good', 'Attack_Type': '', 'Notes': ''},
        {'Set': 1, 'Point': 6, 'Rotation': 3, 'Player': 'Alex', 'Position': 'OH1', 'Action': 'attack', 'Outcome': 'kill', 'Attack_Type': 'normal', 'Notes': ''},
        
        # Set 1, Point 7 - Block kill
        {'Set': 1, 'Point': 7, 'Rotation': 4, 'Player': 'Emma', 'Position': 'MB2', 'Action': 'block', 'Outcome': 'kill', 'Attack_Type': '', 'Notes': ''},
        
        # Set 1, Point 8 - Attack errors
        {'Set': 1, 'Point': 8, 'Rotation': 4, 'Player': 'David', 'Position': 'MB1', 'Action': 'attack', 'Outcome': 'out', 'Attack_Type': 'normal', 'Notes': ''},
        
        # Set 1, Point 9 - Attack into net
        {'Set': 1, 'Point': 9, 'Rotation': 5, 'Player': 'Sarah', 'Position': 'OH2', 'Action': 'attack', 'Outcome': 'net', 'Attack_Type': 'tip', 'Notes': ''},
        
        # Set 1, Point 10 - Reception error
        {'Set': 1, 'Point': 10, 'Rotation': 5, 'Player': 'John', 'Position': 'L', 'Action': 'receive', 'Outcome': 'error', 'Attack_Type': '', 'Notes': ''},
    ]
    
    df_individual = pd.DataFrame(individual_events)
    
    # Sample Team Events
    team_events = [
        {'Set': 1, 'Point': 1, 'Rotation': 1, 'Point_Type': 'serving', 'Point Won': 'yes', 'Our_Score': 1, 'Opponent_Score': 0, 'Rally_Length': 4},
        {'Set': 1, 'Point': 2, 'Rotation': 1, 'Point_Type': 'receiving', 'Point Won': 'no', 'Our_Score': 1, 'Opponent_Score': 1, 'Rally_Length': 3},
        {'Set': 1, 'Point': 3, 'Rotation': 2, 'Point_Type': 'receiving', 'Point Won': 'yes', 'Our_Score': 2, 'Opponent_Score': 1, 'Rally_Length': 3},
        {'Set': 1, 'Point': 4, 'Rotation': 2, 'Point_Type': 'serving', 'Point Won': 'yes', 'Our_Score': 3, 'Opponent_Score': 1, 'Rally_Length': 5},
        {'Set': 1, 'Point': 5, 'Rotation': 3, 'Point_Type': 'serving', 'Point Won': 'yes', 'Our_Score': 4, 'Opponent_Score': 1, 'Rally_Length': 1},
        {'Set': 1, 'Point': 6, 'Rotation': 3, 'Point_Type': 'serving', 'Point Won': 'yes', 'Our_Score': 5, 'Opponent_Score': 1, 'Rally_Length': 7},
        {'Set': 1, 'Point': 7, 'Rotation': 4, 'Point_Type': 'serving', 'Point Won': 'yes', 'Our_Score': 6, 'Opponent_Score': 1, 'Rally_Length': 1},
        {'Set': 1, 'Point': 8, 'Rotation': 4, 'Point_Type': 'serving', 'Point Won': 'no', 'Our_Score': 6, 'Opponent_Score': 2, 'Rally_Length': 1},
        {'Set': 1, 'Point': 9, 'Rotation': 5, 'Point_Type': 'receiving', 'Point Won': 'no', 'Our_Score': 6, 'Opponent_Score': 3, 'Rally_Length': 1},
        {'Set': 1, 'Point': 10, 'Rotation': 5, 'Point_Type': 'receiving', 'Point Won': 'no', 'Our_Score': 6, 'Opponent_Score': 4, 'Rally_Length': 1},
    ]
    
    df_team = pd.DataFrame(team_events)
    
    # Create Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_individual.to_excel(writer, sheet_name='Individual Events', index=False)
        df_team.to_excel(writer, sheet_name='Team Events', index=False)
        
        # Format headers
        from openpyxl.styles import Font, PatternFill
        
        header_fill = PatternFill(start_color="040C7B", end_color="040C7B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for ws in [writer.sheets['Individual Events'], writer.sheets['Team Events']]:
            for cell in ws[1]:  # Header row
                cell.font = header_font
                cell.fill = header_fill
        
        # Auto-adjust column widths
        for ws in [writer.sheets['Individual Events'], writer.sheets['Team Events']]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Sample event tracker data created at: {output_path}")
    print(f"   - Individual Events: {len(df_individual)} events")
    print(f"   - Team Events: {len(df_team)} points")
    return output_path


if __name__ == "__main__":
    import sys
    output = sys.argv[1] if len(sys.argv) > 1 else "../data/examples/sample_event_tracker.xlsx"
    create_sample_event_data(output)

