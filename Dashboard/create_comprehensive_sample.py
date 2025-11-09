"""
Create comprehensive sample event tracker data for a 3-0 match
With realistic volleyball logic: proper rotations, score tracking, rally sequences
"""
import pandas as pd
import random
from pathlib import Path
from typing import List, Dict, Tuple, Optional

# Set random seed for reproducibility (changed to get more diverse attack types)
random.seed(123)

def get_next_rotation(current_rotation: int, we_served: bool, we_won: bool) -> int:
    """Calculate next rotation based on volleyball rules.
    
    CRITICAL: Rotations MUST experience both serving and receiving before moving to next rotation.
    Rotations go counterclockwise: 1 → 6 → 5 → 4 → 3 → 2 → 1
    
    Rotation changes ONLY when:
    - We win while receiving → rotate counterclockwise and start serving
    
    Rotation stays same when:
    - We win while serving → keep serving (same rotation)
    - We lose while serving → start receiving (same rotation)
    - We lose while receiving → keep receiving (same rotation)
    """
    if we_served and we_won:
        # We keep serving, rotation stays same
        return current_rotation
    elif we_served and not we_won:
        # We lose serve, rotation stays same (we'll be receiving)
        return current_rotation
    elif not we_served and we_won:
        # We gain serve by winning while receiving, rotate counterclockwise
        # Counterclockwise: 1→6, 6→5, 5→4, 4→3, 3→2, 2→1
        if current_rotation == 1:
            return 6
        elif current_rotation == 6:
            return 5
        elif current_rotation == 5:
            return 4
        elif current_rotation == 4:
            return 3
        elif current_rotation == 3:
            return 2
        else:  # current_rotation == 2
            return 1
    else:  # not we_served and not we_won
        # Opponent keeps serving, rotation stays same (we keep receiving)
        return current_rotation

def get_player_for_position(position: str, rotation: int, players_in_set: List[str], 
                           positions_map: Dict[str, str], libero: Optional[str] = None) -> str:
    """Get player name for a position in a given rotation.
    
    In volleyball, positions are fixed relative to rotation:
    Rotation 1: Server (position varies)
    Rotation 2-6: Follow clockwise
    
    For simplicity, we'll use a mapping that rotates players through positions.
    """
    # Simple rotation-based player assignment
    # In real volleyball, players rotate through positions, but for sample data
    # we'll use a simplified approach where we select from available players
    available_players = [p for p in players_in_set if positions_map[p] == position]
    
    # If libero is available and position is L (back row), use libero
    if position == 'L' and libero:
        return libero
    
    # For other positions, rotate through available players
    if available_players:
        # Use rotation to pick player
        idx = (rotation - 1) % len(available_players)
        return available_players[idx]
    
    # Fallback: return first player that matches position
    for player in players_in_set:
        if positions_map[player] == position:
            return player
    
    return players_in_set[0]  # Fallback

def create_rally_sequence(point_type: str, we_win: bool, rally_length: int, 
                         rotation: int, players_in_set: List[str], 
                         positions_map: Dict[str, str], setter: str, libero: Optional[str]) -> List[Dict]:
    """Create a realistic rally sequence of actions."""
    events = []
    
    if point_type == 'serving':
        # We serve
        server_position = ['OH1', 'OH2', 'OPP', 'MB1', 'MB2', 'S'][(rotation - 1) % 6]
        server = get_player_for_position(server_position, rotation, players_in_set, positions_map)
        
        if rally_length == 1:
            # Ace or service error
            outcome = 'ace' if we_win else 'error'
            events.append({
                'Player': server, 'Position': server_position, 'Action': 'serve', 
                'Outcome': outcome, 'Attack_Type': '', 'Notes': ''
            })
        else:
            # Normal serve
            events.append({
                'Player': server, 'Position': server_position, 'Action': 'serve', 
                'Outcome': 'good', 'Attack_Type': '', 'Notes': ''
            })
            
            # Opponent receives, sets, and attacks (we don't track opponent actions)
            # Our defensive response starts here
            
            if rally_length >= 2:
                # Block attempt (opponent attacks, we block)
                blocker_position = ['MB1', 'MB2'][(rotation - 1) % 2]
                blocker = get_player_for_position(blocker_position, rotation, players_in_set, positions_map)
                block_outcome = random.choices(['kill', 'touch', 'missed', 'error'], weights=[0.1, 0.6, 0.2, 0.1])[0]
                if random.random() < 0.5:  # 50% chance of block attempt
                    events.append({
                        'Player': blocker, 'Position': blocker_position, 'Action': 'block', 
                        'Outcome': block_outcome, 'Attack_Type': '', 'Notes': ''
                    })
                    
                    if block_outcome == 'kill':
                        # Block kill ends rally
                        return events
            
            # Our dig (if block touched or attack came through)
            if rally_length >= 3 and libero:
                dig_quality = random.choices(['perfect', 'good', 'poor'], weights=[0.4, 0.5, 0.1])[0]
                events.append({
                    'Player': libero, 'Position': 'L', 'Action': 'dig', 
                    'Outcome': dig_quality, 'Attack_Type': '', 'Notes': ''
                })
            
            # Setter sets
            if rally_length >= 3:
                set_quality = random.choices(['exceptional', 'good', 'poor'], weights=[0.3, 0.6, 0.1])[0]
                events.append({
                    'Player': setter, 'Position': 'S', 'Action': 'set', 
                    'Outcome': set_quality, 'Attack_Type': '', 'Notes': ''
                })
            
            # Attack - prioritize: Outside > Opposite > Middle (more realistic)
            if rally_length >= 4:
                # Weighted selection: 40% OH1, 30% OH2, 20% OPP, 5% MB1, 5% MB2
                attacker_positions = ['OH1', 'OH2', 'OH1', 'OH2', 'OPP', 'OPP', 'MB1', 'MB2']
                attacker_pos = random.choice(attacker_positions)
                attacker = get_player_for_position(attacker_pos, rotation, players_in_set, positions_map, libero)
                
                if we_win:
                    attack_outcome = random.choices(['kill', 'kill', 'defended'], weights=[0.7, 0.2, 0.1])[0]
                else:
                    attack_outcome = random.choices(['blocked', 'out', 'net', 'error'], weights=[0.4, 0.3, 0.2, 0.1])[0]
                
                attack_type = random.choices(['normal', 'tip', 'after_block'], weights=[0.75, 0.15, 0.10])[0]
                events.append({
                    'Player': attacker, 'Position': attacker_pos, 'Action': 'attack', 
                    'Outcome': attack_outcome, 'Attack_Type': attack_type, 'Notes': ''
                })
                
                # If attack is defended/blocked, continue rally (longer rallies)
                if attack_outcome == 'defended' and rally_length >= 5:
                    # Opponent digs and attacks back (we don't track)
                    # Our dig
                    if libero and rally_length >= 5:
                        events.append({
                            'Player': libero, 'Position': 'L', 'Action': 'dig', 
                            'Outcome': random.choice(['perfect', 'good']), 'Attack_Type': '', 'Notes': ''
                        })
                    
                    # Set again
                    if rally_length >= 6:
                        events.append({
                            'Player': setter, 'Position': 'S', 'Action': 'set', 
                            'Outcome': 'good', 'Attack_Type': '', 'Notes': ''
                        })
                    
                    # Attack again - prioritize outside hitters
                    if rally_length >= 7:
                        attacker_positions = ['OH1', 'OH2', 'OH1', 'OH2', 'OPP', 'OPP', 'MB1', 'MB2']
                        attacker_pos = random.choice(attacker_positions)
                        attacker = get_player_for_position(attacker_pos, rotation, players_in_set, positions_map, libero)
                        events.append({
                            'Player': attacker, 'Position': attacker_pos, 'Action': 'attack', 
                            'Outcome': 'kill' if we_win else random.choice(['blocked', 'out']), 
                            'Attack_Type': 'after_block', 'Notes': ''
                        })
    
    else:  # receiving
        # We receive - balance between libero and outside hitters (more realistic)
        # Libero receives ~60% of time, outside hitters ~40%
        if libero and random.random() < 0.6:
            # Libero receives
            receive_quality = random.choices(['perfect', 'good', 'poor', 'error'], weights=[0.35, 0.5, 0.12, 0.03])[0]
            events.append({
                'Player': libero, 'Position': 'L', 'Action': 'receive', 
                'Outcome': receive_quality, 'Attack_Type': '', 'Notes': ''
            })
        else:
            # Outside hitter receives (more realistic)
            # Choose between OH1 and OH2
            oh_positions = ['OH1', 'OH2']
            receiver_pos = random.choice(oh_positions)
            receiver = get_player_for_position(receiver_pos, rotation, players_in_set, positions_map)
            # Outside hitters slightly lower quality than libero
            receive_quality = random.choices(['perfect', 'good', 'poor', 'error'], weights=[0.25, 0.5, 0.2, 0.05])[0]
            events.append({
                'Player': receiver, 'Position': receiver_pos, 'Action': 'receive', 
                'Outcome': receive_quality, 'Attack_Type': '', 'Notes': ''
            })
        
        if receive_quality == 'error':
            return events  # Reception error ends rally
        
        # Setter sets
        if rally_length >= 2:
            set_quality = 'exceptional' if events[0]['Outcome'] == 'perfect' else random.choice(['exceptional', 'good', 'poor'])
            events.append({
                'Player': setter, 'Position': 'S', 'Action': 'set', 
                'Outcome': set_quality, 'Attack_Type': '', 'Notes': ''
            })
        
        # Attack - prioritize: Outside > Opposite > Middle (more realistic)
        if rally_length >= 3:
            # Weighted selection: 40% OH1, 30% OH2, 20% OPP, 5% MB1, 5% MB2
            attacker_positions = ['OH1', 'OH2', 'OH1', 'OH2', 'OPP', 'OPP', 'MB1', 'MB2']
            attacker_pos = random.choice(attacker_positions)
            attacker = get_player_for_position(attacker_pos, rotation, players_in_set, positions_map, libero)
            
            if we_win:
                attack_outcome = random.choices(['kill', 'kill', 'defended'], weights=[0.75, 0.2, 0.05])[0]
            else:
                attack_outcome = random.choices(['blocked', 'out', 'net', 'error'], weights=[0.4, 0.3, 0.2, 0.1])[0]
            
            attack_type = random.choice(['normal', 'tip', 'after_block'])
            events.append({
                'Player': attacker, 'Position': attacker_pos, 'Action': 'attack', 
                'Outcome': attack_outcome, 'Attack_Type': attack_type, 'Notes': ''
            })
            
            # If attack is defended/blocked, continue rally (longer rallies)
            if attack_outcome in ['defended', 'blocked'] and rally_length >= 4:
                # Dig (if attack was defended)
                if attack_outcome == 'defended' and libero and rally_length >= 4:
                    events.append({
                        'Player': libero, 'Position': 'L', 'Action': 'dig', 
                        'Outcome': random.choice(['perfect', 'good']), 'Attack_Type': '', 'Notes': ''
                    })
                
                # Set again
                if rally_length >= 5:
                    events.append({
                        'Player': setter, 'Position': 'S', 'Action': 'set', 
                        'Outcome': 'good', 'Attack_Type': '', 'Notes': ''
                    })
                
                # Attack again - prioritize outside hitters
                if rally_length >= 6:
                    attacker_positions = ['OH1', 'OH2', 'OH1', 'OH2', 'OPP', 'OPP', 'MB1', 'MB2']
                    attacker_pos = random.choice(attacker_positions)
                    attacker = get_player_for_position(attacker_pos, rotation, players_in_set, positions_map, libero)
                    second_attack_outcome = 'kill' if we_win else random.choice(['blocked', 'out', 'defended'])
                    events.append({
                        'Player': attacker, 'Position': attacker_pos, 'Action': 'attack', 
                        'Outcome': second_attack_outcome, 
                        'Attack_Type': 'after_block', 'Notes': ''
                    })
                    
                    # Third cycle if needed (very long rallies)
                    if second_attack_outcome == 'defended' and rally_length >= 7:
                        if libero:
                            events.append({
                                'Player': libero, 'Position': 'L', 'Action': 'dig', 
                                'Outcome': random.choice(['perfect', 'good']), 'Attack_Type': '', 'Notes': ''
                            })
                        
                        if rally_length >= 8:
                            events.append({
                                'Player': setter, 'Position': 'S', 'Action': 'set', 
                                'Outcome': 'good', 'Attack_Type': '', 'Notes': ''
                            })
                            
                            attacker_positions = ['OH1', 'OH2', 'OH1', 'OH2', 'OPP', 'OPP', 'MB1', 'MB2']
                            attacker_pos = random.choice(attacker_positions)
                            attacker = get_player_for_position(attacker_pos, rotation, players_in_set, positions_map, libero)
                            events.append({
                                'Player': attacker, 'Position': attacker_pos, 'Action': 'attack', 
                                'Outcome': 'kill' if we_win else 'blocked', 
                                'Attack_Type': 'after_block', 'Notes': ''
                            })
    
    return events

def generate_set(set_num: int, target_score: Tuple[int, int], players_in_set: List[str], 
                positions_map: Dict[str, str], setter: str, libero: Optional[str]) -> Tuple[List[Dict], List[Dict]]:
    """Generate a complete set with realistic volleyball logic.
    
    Ensures each rotation experiences both serving and receiving opportunities.
    """
    individual_events = []
    team_events = []
    
    our_score = 0
    opp_score = 0
    point = 1
    rotation = 1
    we_are_serving = True  # Start serving
    
    target_our, target_opp = target_score
    
    # Track rotation usage to ensure balance
    rotation_serving_count = {r: 0 for r in range(1, 7)}
    rotation_receiving_count = {r: 0 for r in range(1, 7)}
    
    # Track consecutive serves/receives to force exchanges
    consecutive_serves = 0
    consecutive_receives = 0
    
    # Track complete rotation cycles (1→2→3→4→5→6→1 = 1 complete cycle)
    rotation_cycle_count = 0
    last_rotation_in_cycle = None
    
    # Track if we've reached the target score
    set_complete = False
    
    while not set_complete:
        # Check if set is complete (both teams have reached or exceeded targets)
        if our_score >= target_our and opp_score >= target_opp:
            set_complete = True
            break
        
        # Determine if we win this point
        points_remaining_our = max(0, target_our - our_score)
        points_remaining_opp = max(0, target_opp - opp_score)
        
        # If we've reached our target but opponent hasn't, opponent wins remaining
        if our_score >= target_our and opp_score < target_opp:
            we_win = False
        # If opponent has reached target but we haven't, we win remaining
        elif opp_score >= target_opp and our_score < target_our:
            we_win = True
        # If both still need points, calculate probability
        elif points_remaining_our > 0 and points_remaining_opp > 0:
            total_remaining = points_remaining_our + points_remaining_opp
            win_probability = points_remaining_our / total_remaining
            # Keep win probability closer to 50/50 to ensure rotation cycling
            win_probability = max(0.4, min(0.7, win_probability))
            
            # CRITICAL: Force serve exchanges to ensure rotation cycling
            # The key insight: Rotations alternate (1,3,5 serve; 2,4,6 receive)
            # To break this pattern, we need rotations to sometimes WIN while receiving
            # This allows them to serve in the next rotation, breaking the alternating pattern
            base_win_prob = 0.5  # Default 50/50
            
            # Check if rotations need balancing - use CURRENT state
            rotations_need_serving = [r for r in range(1, 7) if rotation_receiving_count[r] > 0 and rotation_serving_count[r] == 0]
            rotations_need_receiving = [r for r in range(1, 7) if rotation_serving_count[r] > 0 and rotation_receiving_count[r] == 0]
            
            # Count how many rotations have experienced both
            rotations_with_both = sum(1 for r in range(1, 7) 
                                     if rotation_serving_count[r] > 0 and rotation_receiving_count[r] > 0)
            
            # CRITICAL: To break the alternating pattern, we need rotations 2,4,6 to sometimes WIN while receiving
            # This allows them to serve in the next rotation
            if we_are_serving:
                # If current rotation has served but never received, force serve loss
                if rotation_serving_count[rotation] > 0 and rotation_receiving_count[rotation] == 0:
                    base_win_prob = 0.01  # 99% chance of losing serve
                # If any rotation needs receiving, increase chance of serve loss
                elif len(rotations_need_receiving) > 0:
                    base_win_prob = 0.1  # 90% chance of losing serve
                # Otherwise, allow some serving streaks but not too long
                elif consecutive_serves >= 2:
                    base_win_prob = 0.3
                else:
                    base_win_prob = 0.5
            else:
                # CRITICAL: To break the alternating pattern, we need rotations to WIN while receiving
                # This allows them to serve in the next rotation, which breaks the pattern
                # The key: when rotation 2 receives and wins → rotation 3 serves
                # But for rotation 2 to serve, rotation 1 must receive and win → rotation 2 serves
                # So we need to ensure rotations cycle through multiple times
                
                # Count how many rotations have experienced both
                rotations_with_both = sum(1 for r in range(1, 7) 
                                         if rotation_serving_count[r] > 0 and rotation_receiving_count[r] > 0)
                
                # CRITICAL: If ANY rotation hasn't experienced both, ALWAYS win while receiving
                # This forces rotations to cycle through multiple times, ensuring each experiences both
                # The key insight: winning while receiving breaks the alternating pattern
                # Also check rotation cycles - need at least 2 complete cycles for all rotations to experience both
                if rotations_with_both < 6 or rotation_cycle_count < 2:
                    # ABSOLUTE priority: win while receiving to force rotation cycling
                    # This is the ONLY way to break the alternating pattern
                    base_win_prob = 0.99  # 99% chance - almost certain
                # If this rotation has never served, win to allow it to serve next
                elif rotation_serving_count[rotation] == 0:
                    base_win_prob = 0.95  # 95% chance of gaining serve
                # If this rotation has received but never served, win to allow it to serve
                elif rotation_receiving_count[rotation] > 0 and rotation_serving_count[rotation] == 0:
                    base_win_prob = 0.99  # 99% chance of gaining serve
                # If any rotation needs serving, increase chance of serve gain
                elif len(rotations_need_serving) > 0:
                    base_win_prob = 0.9  # 90% chance of gaining serve
                # Otherwise, allow some receiving streaks but not too long
                elif consecutive_receives >= 2:
                    base_win_prob = 0.7
                else:
                    base_win_prob = 0.5
            
            # Blend with target-based probability (rotation balancing takes ABSOLUTE priority)
            win_probability = (base_win_prob * 0.98) + (win_probability * 0.02)
            
            we_win = random.random() < win_probability
        else:
            # Fallback
            we_win = our_score < target_our
        
        # Determine rally length (3-5 average, some longer)
        if random.random() < 0.15:  # 15% long rallies
            rally_length = random.randint(5, 8)
        elif random.random() < 0.1:  # 10% very short (ace/error)
            rally_length = 1
        else:  # 75% normal rallies
            rally_length = random.randint(2, 5)
        
        # Create rally sequence
        point_type = 'serving' if we_are_serving else 'receiving'
        rally_events = create_rally_sequence(point_type, we_win, rally_length, rotation, 
                                            players_in_set, positions_map, setter, libero)
        
        # Track rotation usage
        if point_type == 'serving':
            rotation_serving_count[rotation] += 1
        else:
            rotation_receiving_count[rotation] += 1
        
        # Add set, point, rotation to each event
        for event in rally_events:
            event['Set'] = set_num
            event['Point'] = point
            event['Rotation'] = rotation
            individual_events.append(event.copy())
        
        # Update scores
        if we_win:
            our_score += 1
        else:
            opp_score += 1
        
        # Add team event
        team_events.append({
            'Set': set_num,
            'Point': point,
            'Rotation': rotation,
            'Point_Type': point_type,
            'Point Won': 'yes' if we_win else 'no',
            'Our_Score': our_score,
            'Opponent_Score': opp_score,
            'Rally_Length': len(rally_events)
        })
        
        # Update rotation and serving status
        old_rotation = rotation
        rotation = get_next_rotation(rotation, we_are_serving, we_win)
        
        # Track rotation cycles (when we complete a full cycle 1→6→5→4→3→2→1)
        if last_rotation_in_cycle is not None:
            # Check if we've completed a cycle (went from rotation 2 back to rotation 1, counterclockwise)
            if old_rotation == 2 and rotation == 1:
                rotation_cycle_count += 1
        last_rotation_in_cycle = rotation
        
        # Update serving status and track consecutive serves/receives
        if we_win and not we_are_serving:
            # We gained serve by winning while receiving - rotation changed, now serving
            we_are_serving = True
            consecutive_receives = 0
            consecutive_serves = 1
        elif not we_win and we_are_serving:
            # We lost serve - rotation stays same, now receiving
            we_are_serving = False
            consecutive_serves = 0
            consecutive_receives = 1
        elif we_are_serving:
            # We kept serve - rotation stays same
            consecutive_serves += 1
            consecutive_receives = 0
        else:
            # We kept receiving - rotation stays same
            consecutive_receives += 1
            consecutive_serves = 0
        
        point += 1
        
        # Safety break
        if point > 100:
            break
    
    return individual_events, team_events

def create_comprehensive_sample(output_path: str = "../data/examples/comprehensive_match_3_0.xlsx"):
    """Create comprehensive sample event tracker data for a 3-0 win"""
    
    # Create directory if it doesn't exist
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Player positions mapping
    positions_map = {
        'Grzegorz': 'S',   # Setter - all 3 sets
        'Luka': 'MB1',     # Middle - all 3 sets
        'Fabio': 'MB2',    # Middle - all 3 sets
        'Stefano': 'OH1',  # Outside - Set 1, 2
        'Sohel': 'OH2',    # Outside - Set 1, 3
        'Vincent': 'OPP',  # Opposite - Set 1, 2
        'Alex': 'L',       # Libero - Set 1, 3
        'Mladen': 'OH1',   # Outside - Set 2, 3
        'Mariusz': 'OPP',  # Opposite - Set 3
        'David': 'L'       # Libero - Set 2
    }
    
    all_individual_events = []
    all_team_events = []
    
    # Set 1: 25-20 (we win)
    # Players: Grzegorz, Luka, Fabio, Stefano, Sohel, Vincent, Alex
    set1_players = ['Grzegorz', 'Luka', 'Fabio', 'Stefano', 'Sohel', 'Vincent', 'Alex']
    ind1, team1 = generate_set(1, (25, 20), set1_players, positions_map, 'Grzegorz', 'Alex')
    all_individual_events.extend(ind1)
    all_team_events.extend(team1)
    
    # Set 2: 25-18 (we win)
    # Players: Grzegorz, Luka, Fabio, Stefano, Mladen, Vincent, David
    set2_players = ['Grzegorz', 'Luka', 'Fabio', 'Stefano', 'Mladen', 'Vincent', 'David']
    ind2, team2 = generate_set(2, (25, 18), set2_players, positions_map, 'Grzegorz', 'David')
    all_individual_events.extend(ind2)
    all_team_events.extend(team2)
    
    # Set 3: 25-22 (we win)
    # Players: Grzegorz, Luka, Fabio, Sohel, Mladen, Mariusz, Alex
    set3_players = ['Grzegorz', 'Luka', 'Fabio', 'Sohel', 'Mladen', 'Mariusz', 'Alex']
    ind3, team3 = generate_set(3, (25, 22), set3_players, positions_map, 'Grzegorz', 'Alex')
    all_individual_events.extend(ind3)
    all_team_events.extend(team3)
    
    # Create DataFrames
    df_individual = pd.DataFrame(all_individual_events)
    df_team = pd.DataFrame(all_team_events)
    
    # Ensure proper column order
    individual_columns = ['Set', 'Point', 'Rotation', 'Player', 'Position', 'Action', 'Outcome', 'Attack_Type', 'Notes']
    team_columns = ['Set', 'Point', 'Rotation', 'Point_Type', 'Point Won', 'Our_Score', 'Opponent_Score', 'Rally_Length']
    
    df_individual = df_individual[individual_columns]
    df_team = df_team[team_columns]
    
    # Create Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_individual.to_excel(writer, sheet_name='Individual Events', index=False)
        df_team.to_excel(writer, sheet_name='Team Events', index=False)
        
        # Format headers
        from openpyxl.styles import Font, PatternFill
        
        header_fill = PatternFill(start_color="040C7B", end_color="040C7B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for ws in [writer.sheets['Individual Events'], writer.sheets['Team Events']]:
            for cell in ws[1]:  # Header row
                cell.font = header_font
                cell.fill = header_fill
        
        # Auto-adjust column widths
        for ws in [writer.sheets['Individual Events'], writer.sheets['Team Events']]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ Comprehensive match data created at: {output_path}")
    print(f"   - Individual Events: {len(df_individual)} events")
    print(f"   - Team Events: {len(df_team)} points")
    
    # Quality checks
    print("\n📊 Quality Checks:")
    
    # Verify scores
    for set_num in [1, 2, 3]:
        set_data = df_team[df_team['Set'] == set_num]
        final_our = set_data['Our_Score'].max()
        final_opp = set_data['Opponent_Score'].max()
        print(f"   ✓ Set {set_num}: {final_our}-{final_opp}")
    
    # Check rotation balance
    print("\n   Rotation Balance (Serving/Receiving):")
    all_rotations_ok = True
    for set_num in [1, 2, 3]:
        set_data = df_team[df_team['Set'] == set_num]
        print(f"   Set {set_num}:")
        for rot in sorted(set_data['Rotation'].unique()):
            rot_data = set_data[set_data['Rotation'] == rot]
            serving = len(rot_data[rot_data['Point_Type'] == 'serving'])
            receiving = len(rot_data[rot_data['Point_Type'] == 'receiving'])
            
            if serving == 0 or receiving == 0:
                all_rotations_ok = False
                print(f"     ⚠️  Rotation {rot}: {serving} serving, {receiving} receiving (UNBALANCED)")
            else:
                print(f"     ✓ Rotation {rot}: {serving} serving, {receiving} receiving")
    
    # Check action distribution
    print("\n   Action Distribution:")
    action_counts = df_individual['Action'].value_counts()
    for action, count in action_counts.items():
        pct = (count / len(df_individual)) * 100
        print(f"     {action.capitalize()}: {count} ({pct:.1f}%)")
    
    # Check attack type distribution
    attacks = df_individual[df_individual['Action'] == 'attack']
    if len(attacks) > 0:
        print("\n   Attack Type Distribution:")
        attack_types = attacks['Attack_Type'].value_counts()
        for atype, count in attack_types.items():
            if atype and str(atype).strip():
                pct = (count / len(attacks)) * 100
                print(f"     {atype.capitalize()}: {count} ({pct:.1f}%)")
    
    # Check outcome distribution
    print("\n   Outcome Distribution (Top 10):")
    outcome_counts = df_individual['Outcome'].value_counts().head(10)
    for outcome, count in outcome_counts.items():
        pct = (count / len(df_individual)) * 100
        print(f"     {outcome.capitalize()}: {count} ({pct:.1f}%)")
    
    # Check reception distribution (should be mostly libero and outside hitters)
    receives = df_individual[df_individual['Action'] == 'receive']
    if len(receives) > 0:
        print("\n   Reception by Position:")
        rec_by_pos = receives['Position'].value_counts()
        for pos, count in rec_by_pos.items():
            pct = (count / len(receives)) * 100
            print(f"     {pos}: {count} ({pct:.1f}%)")
    
    # Check attack distribution (should favor outside hitters)
    if len(attacks) > 0:
        print("\n   Attack by Position:")
        attack_by_pos = attacks['Position'].value_counts()
        for pos, count in attack_by_pos.items():
            pct = (count / len(attacks)) * 100
            print(f"     {pos}: {count} ({pct:.1f}%)")
    
    if all_rotations_ok:
        print("\n✅ All quality checks passed!")
    else:
        print("\n⚠️  Some rotations are unbalanced - please review the data")
        print("   Note: In very short sets, some rotations may not experience both serving and receiving")
    
    return output_path


if __name__ == "__main__":
    import sys
    output = sys.argv[1] if len(sys.argv) > 1 else "../data/examples/comprehensive_match_3_0.xlsx"
    create_comprehensive_sample(output)
